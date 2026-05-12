#!/usr/bin/env python3
"""
ver2 엑셀 빌더 — 11컬럼, 행 단위 이미지 임베드 (병합 없음)

스킬 샘플 양식 베이스에 K(프로필 스크린샷) 컬럼 추가.
스레드 단위 J 병합 대신 각 행 = 댓글/프로필 스크린샷 1쌍.

J:  output/{POST}/{i}번스레드/{j:02d}.png  (그 댓글의 페이지 캡처)
K:  output/{POST}/profiles/{username}.png  (그 계정 프로필 페이지 캡처)

Usage:
  build_excel.py [progress.json] [out.xlsx]
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.units import pixels_to_EMU

PROJECT = Path(__file__).resolve().parent
DEFAULT_PROGRESS = PROJECT / "output" / ".progress_dxqb.json"

SHEET_NAME = "증거자료"
HEADERS = ["번호", "구분", "계정", "본명", "댓글 내용",
           "댓글 URL", "프로필 URL", "작성 시간", "좋아요",
           "댓글 스크린샷", "프로필 스크린샷"]
COL_WIDTHS = {"A": 7, "B": 10, "C": 22, "D": 18, "E": 60,
              "F": 42, "G": 32, "H": 20, "I": 9,
              "J": 60, "K": 50}
HEADER_ROW_HEIGHT = 28.0
DATA_ROW_HEIGHT = 220.0   # 키워드 행 높이 (이미지가 충분히 보이게)
HEADER_FILL = PatternFill("solid", fgColor="FF3F3F46")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

COMMENT_IMG_COL = 10  # J
PROFILE_IMG_COL = 11  # K
COMMENT_IMG_W = 420
PROFILE_IMG_W = 340


def fmt_display(name, is_private):
    name = (name or "").strip()
    if is_private:
        return f"{name} 🔒".lstrip() if name else " 🔒"
    return name or None


def setup_sheet(ws):
    for letter, w in COL_WIDTHS.items():
        ws.column_dimensions[letter].width = w
    ws.append(HEADERS)
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def write_row(ws, row_idx, num, kind, c):
    ws.cell(row=row_idx, column=1, value=num)
    ws.cell(row=row_idx, column=2, value=kind)
    ws.cell(row=row_idx, column=3, value=c["username"])
    ws.cell(row=row_idx, column=4,
            value=fmt_display(c.get("display_name"), c.get("is_private", False)))
    ws.cell(row=row_idx, column=5, value=c.get("content", ""))
    ws.cell(row=row_idx, column=6, value=c.get("comment_url", ""))
    ws.cell(row=row_idx, column=7, value=c.get("profile_url", ""))
    ws.cell(row=row_idx, column=8, value=c.get("created_at", ""))
    ws.cell(row=row_idx, column=9, value=int(c.get("likes", 0)))

    ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT
    for col in (1, 2, 3, 4, 8, 9, 10, 11):
        ws.cell(row=row_idx, column=col).alignment = Alignment(
            horizontal="center", vertical="center")
    ws.cell(row=row_idx, column=5).alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True)
    for col in (6, 7):
        ws.cell(row=row_idx, column=col).alignment = Alignment(
            horizontal="left", vertical="center")


def embed(ws, col_1based, row, png_path: Path, target_w):
    if not png_path.exists() or png_path.stat().st_size == 0:
        return False
    img = XLImage(str(png_path))
    ratio = target_w / img.width
    new_w = target_w
    new_h = int(img.height * ratio)
    img.width = new_w
    img.height = new_h
    img.anchor = OneCellAnchor(
        _from=AnchorMarker(col=col_1based - 1, colOff=0,
                           row=row - 1, rowOff=0),
        ext=XDRPositiveSize2D(cx=pixels_to_EMU(new_w),
                              cy=pixels_to_EMU(new_h)),
    )
    ws.add_image(img)
    return True


def _safe_name(name):
    import re
    return re.sub(r"[^A-Za-z0-9._\-가-힣]+", "_", name).strip("_") or "unknown"


def build(progress_path: Path, out_xlsx: Path):
    data = json.loads(progress_path.read_text(encoding="utf-8"))
    post_id = data["post_id"]
    folder_name = data.get("folder_name") or post_id  # backward-compat
    total_count = sum(1 + len(t.get("replies", [])) for t in data["threads"])
    # 스크린샷 폴더: output/{folder_name}/스크린샷(N)/
    post_dir = progress_path.resolve().parent / folder_name / f"스크린샷({total_count})"

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    setup_sheet(ws)

    cur = 2
    embedded_comments = 0
    embedded_profiles = 0
    for ti, thread in enumerate(data["threads"], start=1):
        # 위계 표기: 원댓글 = NN_user, 대댓글 = NN_RR_user
        items = [(None, "원댓글", thread["root"])]   # ri=None → root
        for ri, reply in enumerate(thread.get("replies", []), start=1):
            items.append((ri, "대댓글", reply))

        for k, (ri, kind, c) in enumerate(items):
            num = ti if k == 0 else None
            write_row(ws, cur, num, kind, c)

            uname = _safe_name(c["username"])
            if ri is None:
                ent_name = f"{ti:02d}_{uname}"
            else:
                ent_name = f"{ti:02d}_{ri:02d}_{uname}"
            folder = post_dir / ent_name
            comment_png = folder / "댓글.png"
            if embed(ws, COMMENT_IMG_COL, cur, comment_png, COMMENT_IMG_W):
                embedded_comments += 1
            profile_png = folder / "프로필.png"
            if embed(ws, PROFILE_IMG_COL, cur, profile_png, PROFILE_IMG_W):
                embedded_profiles += 1
            cur += 1

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    return out_xlsx, embedded_comments, embedded_profiles, cur - 2


if __name__ == "__main__":
    progress = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_PROGRESS
    if len(sys.argv) > 2:
        out = Path(sys.argv[2])
    else:
        # output/{folder_name}/result.xlsx
        data = json.loads(progress.read_text(encoding="utf-8"))
        folder_name = data.get("folder_name") or data["post_id"]
        out = progress.resolve().parent / folder_name / "result.xlsx"
    saved, ec, ep, rows = build(progress, out)
    print(f"saved: {saved}")
    print(f"data rows: {rows}")
    print(f"embedded comment images: {ec}")
    print(f"embedded profile images: {ep}")
