#!/usr/bin/env python3
"""
이전 xlsx 에서 메타데이터 추출 → ver2 progress.json 생성.

Usage:
  import_meta.py <old_xlsx> <post_id>
"""
import json
import sys
from pathlib import Path
import openpyxl

PROJECT = Path(__file__).resolve().parent


def main():
    if len(sys.argv) != 3:
        print("Usage: import_meta.py <old_xlsx> <post_id>")
        return 1
    old_xlsx = Path(sys.argv[1])
    post_id = sys.argv[2]
    if not old_xlsx.exists():
        print(f"[err] not found: {old_xlsx}")
        return 1

    wb = openpyxl.load_workbook(old_xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]

    threads = []
    cur_thread = None
    for r in range(2, ws.max_row + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, 11)]
        num, kind, uname, bonmyeong, content, curl, purl, ctime, likes, _ = cells
        if not uname:
            continue
        bonmyeong_s = (bonmyeong or "").strip() if bonmyeong else ""
        is_private = "🔒" in bonmyeong_s
        display_name = bonmyeong_s.replace("🔒", "").strip()

        record = {
            "username": uname,
            "display_name": display_name,
            "is_private": is_private,
            "content": content or "",
            "comment_url": curl or "",
            "profile_url": purl or f"https://www.instagram.com/{uname}/",
            "created_at": str(ctime) if ctime else "",
            "likes": int(likes) if likes else 0,
        }

        if kind == "원댓글":
            if cur_thread:
                threads.append(cur_thread)
            cur_thread = {
                "index": len(threads) + 1,
                "root": record,
                "replies": [],
            }
        elif kind == "대댓글":
            if cur_thread:
                cur_thread["replies"].append(record)
            else:
                # 부모 없는 대댓글 — 단독 스레드로
                cur_thread = {"index": len(threads) + 1, "root": record, "replies": []}
                threads.append(cur_thread)
                cur_thread = None
    if cur_thread:
        threads.append(cur_thread)

    # 인덱스 재정리 (1부터 순차)
    for i, t in enumerate(threads, start=1):
        t["index"] = i

    # 산출물 폴더명 (한글 + yymmdd) — 기존 progress 가 있으면 그 안의 folder_name 재사용
    from datetime import datetime, timezone, timedelta
    KST = timezone(timedelta(hours=9))
    existing_folder_name = None
    progress_check = PROJECT / "output" / f".progress_{post_id}.json"
    if progress_check.exists():
        try:
            existing_folder_name = json.loads(progress_check.read_text(encoding="utf-8")).get("folder_name")
        except Exception:
            pass
    date_yymmdd = datetime.now(KST).strftime("%y%m%d")
    folder_name = existing_folder_name or f"인스타_{date_yymmdd}_{post_id}"

    out_data = {
        "post_id": post_id,
        "post_url": f"https://www.instagram.com/p/{post_id}/",
        "phase": "meta_imported",
        "platform": "instagram",
        "platform_label": "인스타",
        "folder_name": folder_name,
        "threads": threads,
    }

    out_path = PROJECT / "output" / f".progress_{post_id}.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(out_data, ensure_ascii=False, indent=2), encoding="utf-8")

    total_comments = sum(1 + len(t["replies"]) for t in threads)
    unique_users = set()
    for t in threads:
        unique_users.add(t["root"]["username"])
        for r in t["replies"]:
            unique_users.add(r["username"])

    print(f"[ok] saved: {out_path}")
    print(f"  threads: {len(threads)}")
    print(f"  total comments (root + replies): {total_comments}")
    print(f"  unique users: {len(unique_users)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
